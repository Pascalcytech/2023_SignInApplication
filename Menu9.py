from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk
import matplotlib.pyplot as plt
import os
import json
import tkinter
import csv
import openpyxl
import datetime


class App(Frame):
    def __init__(self):
        super().__init__()

        # Will contain an instance of the frame_class that is being used to display the current window.
        self.frame = None

        # Will immediately switch to the first window
        self.switch_frame(Menu)

    # Will call other frames with itself as parameter, so they can call this function :
    def switch_frame(self, frame_class):
        if self.frame is not None:
            self.frame.destroy()
        self.frame = frame_class(self)
        self.frame.pack()

    def switch_frame_3arg(self, frame_class, arg1, arg2, arg3):
        if self.frame is not None:
            self.frame.destroy()
        self.frame = frame_class(self, arg1, arg2, arg3)
        self.frame.pack()

# Note à moi-même : lorsque l'on définit une classe (ex Class1), ce qu'on met entre parenthèses n'est pas un paramètre,
# C'est pour indiquer le parent. Quand on appelle la classe en question
# et que ses fonctions n'ont pas besoin de paramètres, on l'appelle sans paramètres : x = Class1()
# Si une de ses fonction a besoin de paramètres (ex : paramètre nom), alors il faut appeller la classe
# avec un paramètre : y = Class1(nom)


class Menu(Frame):
    def __init__(self, parent):
        super().__init__()
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("Application d'émargement")
        # master is a buit-in module : The widget object that contains this widget.
        # calling winfo_parent() returns a string of this widget name whereas master returns the object.

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(5):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        ttk.Button(self, text="Paramètres validation",
                   command=lambda: arg_parent.switch_frame(ValidationParameters)).grid(row=0, column=4)
        ttk.Button(self, text="Importer les élèves",
                   command=lambda: arg_parent.switch_frame(ImportStudents)).grid(row=1, column=2)
        ttk.Button(self, text="Exporter les élèves",
                   command=lambda: arg_parent.switch_frame(ExpAllStudents)).grid(row=2, column=3)
        ttk.Button(self, text="Emargement évènement",
                   command=lambda: arg_parent.switch_frame(SignIn)).grid(row=1, column=3)
        ttk.Button(self, text="Exporter 1 élève",
                   command=lambda: arg_parent.switch_frame(ExpOneStudent)).grid(row=2, column=2)

        # Pie chart for 1 Deparment
        list_of_departments = read_validation_parameters()[1]
        path2 = check_directory() + "\\base de données élèves.csv"
        with open(path2, 'r') as file2:
            csv_reader = csv.reader(file2)
            list_lines = list(csv_reader)

        if len(list_of_departments) > 0 and len(list_lines) > 1:
            name1 = list_of_departments[0][0]  # Taken from the json file
            seuil1 = list_of_departments[0][1]

            indexdname1 = list_lines[0].index("pôle " + name1)  # taken from the csv file
            indexseuil1 = indexdname1 + 1

            totalnbstudent = len(list_lines)-1
            totalvalstudent = 0

            for i in range(1, len(list_lines)):
                if list_lines[i][indexseuil1] >= seuil1:
                    totalvalstudent += 1

            percentage1 = (totalvalstudent / totalnbstudent) * 100

            labels = ["Validé le pôle {}".format(name1), "Non validé"]
            values = [percentage1, 100-percentage1]

            fig, ax = plt.subplots(figsize=(3, 3))
            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
            ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

            canvas = FigureCanvasTkAgg(fig, master=self)
            canvas_widget = canvas.get_tk_widget()
            canvas_widget.grid(row=1, column=0, sticky='nsew')

        if len(list_of_departments) > 1 and len(list_lines) > 1:
            name2 = list_of_departments[1][0]
            seuil2 = list_of_departments[1][1]

            indexdname2 = list_lines[0].index("pôle " + name2)
            indexseuil2 = indexdname2 + 1

            totalnbstudent = len(list_lines) - 1
            totalvalstudent2 = 0

            for i in range(1, len(list_lines)):
                if list_lines[i][indexseuil2] >= seuil2:
                    totalvalstudent2 += 1

            percentage2 = (totalvalstudent2 / totalnbstudent) * 100
            labels2 = ["Validé le pôle {}".format(name2), "Non validé"]
            values2 = [percentage2, 100 - percentage2]

            fig, ax = plt.subplots(figsize=(3, 3))
            ax.pie(values2, labels=labels2, autopct='%1.1f%%', startangle=90)
            ax.axis('equal')

            canvas = FigureCanvasTkAgg(fig, master=self)
            canvas_widget = canvas.get_tk_widget()
            canvas_widget.grid(row=2, column=0, sticky='nsew')

            list_of_actions, list_of_departments = read_validation_parameters()
            text_actions, text_departments = generate_text(list_of_actions, list_of_departments)
            ttk.Label(self, text=text_departments).grid(row=0, column=0)
            ttk.Label(self, text=text_actions).grid(row=0, column=1)


class ValidationParameters(Frame):  # To change the parameters of validation.
    def __init__(self, parent):
        super().__init__()
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("Paramètres de validation")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(3):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        list_of_actions, list_of_departments = read_validation_parameters()
        text_actions, text_departments = generate_text(list_of_actions, list_of_departments)
        ttk.Label(self, text=text_actions).grid(row=0, column=0)
        ttk.Label(self, text=text_departments).grid(row=1, column=0)

        ttk.Button(self, text="Ajouter une action", command=lambda: arg_parent.switch_frame(AddAction)).grid(row=0, column=1)
        ttk.Button(self, text="Ajouter un pôle",
                   command=lambda: arg_parent.switch_frame(AddDepartment)).grid(row=1, column=1)
        ttk.Button(self, text="Retirer une action", command=lambda: arg_parent.switch_frame(DeleteAction)).grid(row=0, column=2)
        ttk.Button(self, text="Retirer un pôle", command=lambda: arg_parent.switch_frame(DeleteDepartment)).grid(row=1, column=2)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(Menu)).grid(row=2, column=2)


class AddDepartment(Frame):
    def __init__(self, parent):
        super().__init__()
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("Ajout d'un pôle")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(3):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        list_of_actions, list_of_departments = read_validation_parameters()
        text_actions, text_departments = generate_text(list_of_actions, list_of_departments)
        ttk.Label(self, text=text_actions).grid(row=0, column=0)
        ttk.Label(self, text=text_departments).grid(row=1, column=0)

        ttk.Label(self, text="Nom du nouveau pôle : ").grid(row=0, column=1)
        ttk.Label(self, text="Nombre de points à atteindre : ").grid(row=1, column=1)
        entry_name = ttk.Entry(self, width=30)
        entry_name.grid(row=0, column=2)
        entry_points = ttk.Entry(self, width=30)
        entry_points.grid(row=1, column=2)

        def confirmation():
            if str(entry_name.get()) != "":
                try:
                    int(entry_points.get())
                except:
                    return

                if int(entry_points.get()) >= 0:
                    file_path = check_directory() + "\\paramètres validation.json"
                    with open(file_path, 'r') as current_file:
                        current_data = json.load(current_file)

                    new_department = {"nom": str(entry_name.get()),
                                      "seuil": str(entry_points.get())}
                    current_data["departments"].append(new_department)

                    with open(file_path, 'w') as current_file:
                        json.dump(current_data, current_file, indent=2)

                    path2 = check_directory() + "\\base de données élèves.csv"
                    with open(path2, 'r') as file2:
                        csv_reader = csv.reader(file2)
                        list_lines = list(csv_reader)
                        list_lines[0].append("pôle " + str(entry_name.get()))
                        list_lines[0].append("total " + str(entry_name.get()))

                        empty_lines = 0  # Sometimes, there are empty lines, why ???
                        for x in range(0, len(list_lines)):
                            if len(list_lines[x]) == 0:
                                empty_lines += 1

                        if len(list_lines)-empty_lines > 1:
                            for y in range(1, len(list_lines)):
                                list_lines[y].append("Non validé")
                                list_lines[y].append("0")

                    with open(path2, 'w', newline='') as file2:
                        csv_writer = csv.writer(file2)
                        csv_writer.writerows(list_lines)

                    arg_parent.switch_frame(ValidationParameters)

        ttk.Button(self, text="Valider", command=lambda: confirmation()).grid(row=2, column=2)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(ValidationParameters)).grid(row=2, column=1)


class DeleteDepartment(Frame):
    def __init__(self, parent):
        super().__init__()
        self.ui(parent)
        self.current_option = ""

    def ui(self, arg_parent):
        self.master.title("Suppression d'un pôle")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(2):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        list_of_actions, list_of_departments = read_validation_parameters()
        text_actions, text_departments = generate_text(list_of_actions, list_of_departments)
        ttk.Label(self, text=text_actions).grid(row=0, column=0)
        ttk.Label(self, text="Sélectionnez le pôle à supprimer avec le menu déroulant").grid(row=0, column=1)
        ttk.Label(self, text=text_departments).grid(row=1, column=0)

        # Preparing the dropdown menu.
        options = []
        for i in range(len(list_of_departments)):
            options.append(list_of_departments[i][0])
        strtemp = tkinter.StringVar()
        dropdown_menu = ttk.Combobox(self, textvariable=strtemp, values=options)
        dropdown_menu.grid(row=1, column=1)

        def on_selection(event):  # Argument non utilisé.
            self.current_option = strtemp.get()

        dropdown_menu.bind("<<ComboboxSelected>>", on_selection)

        # function to call on confirmation.
        def confirmation():
            file_path = check_directory() + "\\paramètres validation.json"

            with open(file_path, 'r') as file:
                data = json.load(file)
                departments = data.get("departments")

                index = 0  # to get the index of the department to delete.
                for department in departments:
                    if department["nom"] == self.current_option:
                        break
                    else:
                        index += 1

                del data["departments"][index]  # data[1] car data[0] correspond aux actions.

            with open(file_path, 'w') as file:
                json.dump(data, file, indent=2)

            path2 = check_directory() + "\\base de données élèves.csv"
            with open(path2, 'r') as file2:
                csv_reader = csv.reader(file2)
                list_lines = list(csv_reader)

                empty_lines = 0  # Sometimes, there are empty lines, why ???
                for x in range(0, len(list_lines)):
                    print()
                    if len(list_lines[x]) == 0:
                        empty_lines += 1

                index1 = list_lines[0].index("pôle " + self.current_option)
                for y in range(0, len(list_lines)-empty_lines):
                    del list_lines[y][index1]

                index2 = list_lines[0].index("total " + self.current_option)
                for z in range(0, len(list_lines)-empty_lines):
                    del list_lines[z][index2]

            with open(path2, 'w', newline='') as file2:
                csv_writer = csv.writer(file2)
                csv_writer.writerows(list_lines)

            # We need to delete the corresponding actions too
            for _ in range (0, len(list_of_actions)):
                if list_of_actions[_][1] == self.current_option:
                    # Name is : list_of_actions[_][0]

                    file_path = check_directory() + "\\paramètres validation.json"

                    with open(file_path, 'r') as file:
                        data = json.load(file)
                        actions = data.get("actions")

                        index = 0  # to get the index of the action to delete.
                        for action in actions:
                            if action["nom"] == list_of_actions[_][0]:
                                break
                            else:
                                index += 1

                        del data["actions"][index]  # data[1] car data[0] correspond aux actions.

                    with open(file_path, 'w') as file:
                        json.dump(data, file, indent=2)

                    path2 = check_directory() + "\\base de données élèves.csv"
                    with open(path2, 'r') as file2:
                        csv_reader = csv.reader(file2)
                        list_lines = list(csv_reader)

                        empty_lines = 0
                        for x in range(0, len(list_lines)):
                            print()
                            if len(list_lines[x]) == 0:
                                empty_lines += 1

                        index1 = list_lines[0].index("total " + list_of_actions[_][0])
                        for y in range(0, len(list_lines) - empty_lines):
                            del list_lines[y][index1]

                    with open(path2, 'w', newline='') as file2:
                        csv_writer = csv.writer(file2)
                        csv_writer.writerows(list_lines)

            arg_parent.switch_frame(ValidationParameters)

        ttk.Button(self, text="Valider", command=lambda: confirmation()).grid(row=2, column=1)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(ValidationParameters)).grid(row=2,
                                                                                                            column=0)


class AddAction(Frame):
    def __init__(self, parent):
        super().__init__()
        self.ui(parent)
        self.current_option = ""

    def ui(self, arg_parent):
        self.master.title("Ajout d'une action")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(3):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        list_of_actions, list_of_departments = read_validation_parameters()
        text_actions, text_departments = generate_text(list_of_actions, list_of_departments)
        ttk.Label(self, text=text_actions).grid(row=0, column=0)
        ttk.Label(self, text=text_departments).grid(row=1, column=0)

        ttk.Label(self, text="Nom de la nouvelle action : ").grid(row=0, column=1)
        ttk.Label(self, text="Sélectionnez le pôle à laquelle l'action appartient avec le menu déroulant").grid(row=1, column=1)
        entry_name = ttk.Entry(self, width=30)
        entry_name.grid(row=0, column=2)

        options = []
        for i in range(len(list_of_departments)):
            options.append(list_of_departments[i][0])
        strtemp = tkinter.StringVar()
        dropdown_menu = ttk.Combobox(self, textvariable=strtemp, values=options)
        dropdown_menu.grid(row=1, column=2)

        def on_selection(event):  # Argument non utilisé.
            self.current_option = strtemp.get()

        dropdown_menu.bind("<<ComboboxSelected>>", on_selection)

        def confirmation():
            if str(entry_name.get()) != "" and self.current_option != "":
                file_path = check_directory() + "\\paramètres validation.json"
                with open(file_path, 'r') as current_file:
                    current_data = json.load(current_file)

                new_action = {"nom": str(entry_name.get()),
                              "pôle": self.current_option}
                current_data["actions"].append(new_action)

                with open(file_path, 'w') as current_file:
                    json.dump(current_data, current_file, indent=2)

                path2 = check_directory() + "\\base de données élèves.csv"
                with open(path2, 'r') as file2:
                    csv_reader = csv.reader(file2)
                    list_lines = list(csv_reader)

                    list_lines[0].append("total " + str(entry_name.get()))

                    empty_lines = 0
                    for x in range(0, len(list_lines)):
                        print()
                        if len(list_lines[x]) == 0:
                            empty_lines += 1

                    if len(list_lines) - empty_lines > 1:
                        for y in range(1, len(list_lines)):
                            list_lines[y].append("0")

                with open(path2, 'w', newline='') as file2:
                    csv_writer = csv.writer(file2)
                    csv_writer.writerows(list_lines)

                arg_parent.switch_frame(ValidationParameters)

        ttk.Button(self, text="Valider", command=lambda: confirmation()).grid(row=2, column=2)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(ValidationParameters)).grid(row=2,
                                                                                                            column=1)


class DeleteAction(Frame):
    def __init__(self, parent):
        super().__init__()
        self.ui(parent)
        self.current_option = ""

    def ui(self, arg_parent):
        self.master.title("Suppression d'une action")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(2):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        list_of_actions, list_of_departments = read_validation_parameters()
        text_actions, text_departments = generate_text(list_of_actions, list_of_departments)
        ttk.Label(self, text=text_actions).grid(row=0, column=0)
        ttk.Label(self, text="Sélectionnez l'action à supprimer avec le menu déroulant").grid(row=0, column=1)
        ttk.Label(self, text=text_departments).grid(row=1, column=0)

        # Preparing the dropdown menu.
        options = []
        for i in range(len(list_of_actions)):
            options.append(list_of_actions[i][0])
        strtemp = tkinter.StringVar()
        dropdown_menu = ttk.Combobox(self, textvariable=strtemp, values=options)
        dropdown_menu.grid(row=1, column=1)

        def on_selection(event):  # Argument non utilisé.
            self.current_option = strtemp.get()

        dropdown_menu.bind("<<ComboboxSelected>>", on_selection)

        # function to call on confirmation.
        def confirmation():
            file_path = check_directory() + "\\paramètres validation.json"

            with open(file_path, 'r') as file:
                data = json.load(file)
                actions = data.get("actions")

                index = 0  # to get the index of the action to delete.
                for action in actions:
                    if action["nom"] == self.current_option:
                        break
                    else:
                        index += 1

                del data["actions"][index]  # data[1] car data[0] correspond aux actions.

            with open(file_path, 'w') as file:
                json.dump(data, file, indent=2)

            path2 = check_directory() + "\\base de données élèves.csv"
            with open(path2, 'r') as file2:
                csv_reader = csv.reader(file2)
                list_lines = list(csv_reader)

                empty_lines = 0
                for x in range(0, len(list_lines)):
                    print()
                    if len(list_lines[x]) == 0:
                        empty_lines += 1

                index1 = list_lines[0].index("total " + self.current_option)
                for y in range(0, len(list_lines) - empty_lines):
                    del list_lines[y][index1]

            with open(path2, 'w', newline='') as file2:
                csv_writer = csv.writer(file2)
                csv_writer.writerows(list_lines)

            arg_parent.switch_frame(ValidationParameters)

        ttk.Button(self, text="Valider", command=lambda: confirmation()).grid(row=2, column=1)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(ValidationParameters)).grid(row=2,
                                                                                                            column=0)


class ImportStudents(Frame):
    def __init__(self, parent):
        super().__init__()
        self.current_path = ""
        self.options = []
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("Importation des élèves")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(1):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        ttk.Label(self, text="sélectionnez le fichier excel des élèves à importer").grid(row=0, column=0)

        def on_selection():
            file_path = filedialog.askopenfilename(title="Select a file",
                                                   filetypes=[("All files", "*.*")])
            self.current_path = file_path
            if os.path.isfile(self.current_path):
                path2 = check_directory() + "\\base de données élèves.csv"
                file_extension = os.path.splitext(self.current_path)[1]
                excel_extensions = ['.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm']
                if file_extension.lower() in excel_extensions:
                    with open(path2, 'r') as file2:
                        csv_reader = csv.reader(file2)
                        list_lines = list(csv_reader)
                        new_lines = [list_lines[0]]

                        workbook = openpyxl.load_workbook(self.current_path)
                        sheet = workbook.active
                        nline = 0
                        for row in sheet.iter_rows():
                            if nline != 0:
                                new_line = []
                                for cell in row:
                                    new_line.append(cell.value)
                                new_lines.append(new_line)
                            nline += 1
                        workbook.close()

                    with open(path2, 'w', newline='') as file2:
                        csv_writer = csv.writer(file2)
                        csv_writer.writerows(new_lines)

                    with open(path2, 'r') as file2:
                        csv_reader = csv.reader(file2)
                        list_lines = list(csv_reader)

                        for a in range(1, len(list_lines)):
                            for b in range(0, len(list_lines[0])):
                                if list_lines[0][b][0:4] == "pôle":
                                    list_lines[a].append("Non validé")
                                elif list_lines[0][b][0:5] == "total":
                                    list_lines[a].append("0")
                                elif list_lines[0][b][0:8] == "présence":
                                    list_lines[a].append("Non présent")

                    with open(path2, 'w', newline='') as file2:
                        csv_writer = csv.writer(file2)
                        csv_writer.writerows(list_lines)

                    arg_parent.switch_frame(Menu)

        ttk.Button(self, text="Sélectionnez un fichier", command=lambda: on_selection()).grid(row=1, column=0)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(Menu)).grid(row=2, column=0)


class SignIn(Frame):
    def __init__(self, parent):
        super().__init__()
        self.current_path = ""
        self.options = []
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("Emargement d'une conférence")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(2):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        ttk.Label(self, text="sélectionnez avec le menu déroulant le chemin vers l'excel d'émargement").grid(row=0, column=0)

        def on_selection():  # Argument non utilisé.
            file_path = filedialog.askopenfilename(title="Select a file", filetypes=[("All files", "*.*")])
            self.current_path = file_path
            if os.path.isfile(self.current_path):
                file_extension = os.path.splitext(self.current_path)[1]
                excel_extensions = ['.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm']
                if file_extension.lower() in excel_extensions:
                    workbook = openpyxl.load_workbook(self.current_path)
                    sheet = workbook.active
                    nline = 0
                    header = []
                    for row in sheet.iter_rows():
                        if nline == 0:
                            for cell in row:
                                header.append(cell.value)
                        nline += 1
                    workbook.close()

                    list_event = []
                    for a in range(0, len(header)):
                        if a > 2:
                            list_event.append(header[a])

                    arg_parent.switch_frame_3arg(RecEvent, self.current_path, list_event, 0)

        ttk.Button(self, text="Sélectionnez un fichier", command=lambda: on_selection()).grid(row=1, column=0)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(Menu)).grid(row=2, column=1)


class RecEvent(Frame):
    def __init__(self, parent, arg_path, arg_list_event, arg_nb_event):
        super().__init__()
        self.current_path = arg_path
        self.success = 0
        self.nb_event = arg_nb_event
        self.action = ""
        self.participating_students = []
        self.list_event = arg_list_event
        self.department = ""
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("prise en compte de l'évènement n°{}".format(self.nb_event))

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(2):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        workbook = openpyxl.load_workbook(self.current_path)
        sheet = workbook.active
        nline = 0
        header = []
        list_lines = []
        for row in sheet.iter_rows():
            newline = []
            if nline == 0:
                for cell in row:
                    header.append(cell.value)
                list_lines.append(header)
            if nline != 0:
                for cell in row:
                    if cell.value is None:
                        newline.append(None)
                    else:
                        newline.append(cell.value)
                list_lines.append(newline)
            nline += 1
        workbook.close()

        # current_event = header[2 + self.nb_event]  # 3 premières colonnes : nom, prénom, n°élève

        for i in range(1, len(list_lines)):
            if list_lines[i][3 + self.nb_event] is not None:
                self.participating_students.append(list_lines[i][2])  # num student in the third column

        ttk.Label(self, text="sélectionnez l'action pour laquelle va compter l'émargement n°{}".format(self.nb_event)).grid(row=0, column=0)
        # Preparing the dropdown menu.
        list_of_actions = read_validation_parameters()[0]
        options = []
        for i in range(len(list_of_actions)):
            options.append(list_of_actions[i][0])
        strtemp = tkinter.StringVar()
        dropdown_menu = ttk.Combobox(self, textvariable=strtemp, values=options)
        dropdown_menu.grid(row=1, column=0)

        def on_selection(event):  # Argument non utilisé.
            self.action = strtemp.get()

        dropdown_menu.bind("<<ComboboxSelected>>", on_selection)

        def confirmation():
            path2 = check_directory() + "\\base de données élèves.csv"
            with open(path2, 'r') as file2:
                new_csv_reader = csv.reader(file2)
                new_list_lines = list(new_csv_reader)

                empty_lines = 0
                for x in range(0, len(new_list_lines)):
                    if len(new_list_lines[x]) == 0:
                        empty_lines += 1

                for y in range(0, len(new_list_lines) - empty_lines):
                    new_list_lines[y].append("")

                new_list_lines[0][len(new_list_lines[0])-1] = self.list_event[self.nb_event]

                index_dep = 0
                for k in list_of_actions:
                    if k[0] == self.action:
                        self.department = k[1]
                        index_dep = new_list_lines[0].index("pôle " + self.department)

                index_act = new_list_lines[0].index("total " + self.action)

                for z in range(1, len(new_list_lines) - empty_lines):  # z is a student line
                    if int(new_list_lines[z][3]) in self.participating_students:  # get the number
                        new_list_lines[z][len(new_list_lines[z])-1] = "Présent"
                        new_list_lines[z][index_dep+1] = str(int(new_list_lines[z][index_dep+1]) + 1)
                        # total is the column next to the title
                        new_list_lines[z][index_act] = str(int(new_list_lines[z][index_act]) + 1)

                        list_of_departments = read_validation_parameters()[1]
                        seuil = 0
                        for l in list_of_departments:
                            if l[0] == self.department:
                                seuil = l[1]
                        if int(new_list_lines[z][index_dep+1]) >= int(seuil):
                            new_list_lines[z][index_dep] = "Validé"
                    else:
                        new_list_lines[z][len(new_list_lines[z]) - 1] = "Absent"

            with open(path2, 'w', newline='') as file2:
                csv_writer = csv.writer(file2)
                csv_writer.writerows(new_list_lines)

            # Go to next event
            self.nb_event += 1
            if self.nb_event == len(self.list_event):
                arg_parent.switch_frame(Menu)
            else:
                arg_parent.switch_frame_3arg(RecEvent, self.current_path, self.list_event, self.nb_event)

        ttk.Button(self, text="Valider", command=lambda: confirmation()).grid(row=2, column=1)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(Menu)).grid(row=2, column=0)


class ExpOneStudent(Frame):
    def __init__(self, parent):
        super().__init__()
        self.current_path = ""
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("Export d'un élève")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(2):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        ttk.Label(self, text="Numéro de l'élève : ").grid(row=0, column=0)
        entry_name = ttk.Entry(self, width=30)
        entry_name.grid(row=0, column=1)

        def on_selection():
            directory_path = filedialog.askdirectory(title="Select a directory")
            self.current_path = directory_path

        def confirmation():
            try:
                int(entry_name.get())
            except:
                return
            if os.path.isdir(self.current_path) and self.current_path != "":
                path2 = check_directory() + "\\base de données élèves.csv"
                with open(path2, 'r') as file2:
                    csv_reader = csv.reader(file2)
                    list_lines = list(csv_reader)
                    list_num = []
                    for _ in range(1, len(list_lines)):
                        list_num.append(list_lines[_][3])
                    if str(entry_name.get()) in list_num:
                        try:
                            line_num = list_num.index(str(entry_name.get())) + 1  # index in list num doe not math index in list lines
                        except:
                            return
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active

                        for col_num in range(1, len(list_lines[0])+1):
                            sheet.cell(row=1, column=col_num, value=list_lines[0][col_num-1])
                            sheet.cell(row=2, column=col_num, value=list_lines[line_num][col_num-1])

                        e_path = self.current_path + '/{}.xlsx'.format(entry_name.get())
                        workbook.save(e_path)
                        workbook.close()
                        arg_parent.switch_frame(Menu)
                    else:
                        pass

        ttk.Button(self, text="Sélectionnez le répertoire d'export", command=lambda: on_selection()).grid(row=1, column=0)
        ttk.Button(self, text="Valider", command=lambda: confirmation()).grid(row=2, column=0)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(Menu)).grid(row=2, column=1)


class ExpAllStudents(Frame):
    def __init__(self, parent):
        super().__init__()
        self.current_path = ""
        self.ui(parent)

    def ui(self, arg_parent):
        self.master.title("Export d'un élève")

        for i in range(3):
            self.grid_rowconfigure(i, weight=1, uniform="equal")
        for j in range(2):
            self.grid_columnconfigure(j, weight=1, uniform="equal")

        def on_selection():
            directory_path = filedialog.askdirectory(title="Select a directory")
            self.current_path = directory_path

        def confirmation():
            if os.path.isdir(self.current_path):
                path2 = check_directory() + "\\base de données élèves.csv"
                with open(path2, 'r') as file2:
                    csv_reader = csv.reader(file2)
                    list_lines = list(csv_reader)

                    workbook = openpyxl.Workbook()
                    sheet = workbook.active

                    for col_num in range(1, len(list_lines[0])+1):  # Index on excel begin at 1. On CSV, begins at 0.
                        sheet.cell(row=1, column=col_num, value=list_lines[0][col_num-1])

                    for row_num in range(1, len(list_lines)+1):
                        for col_num2 in range(1, len(list_lines[0]) + 1):
                            sheet.cell(row=row_num, column=col_num2, value=list_lines[row_num-1][col_num2-1])

                    file_name = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    new_file_name = ""
                    for x in range(0, len(file_name)):
                        if file_name[x] == "-" or file_name[x] == ":":
                            new_file_name += "_"
                        else:
                            new_file_name += file_name[x]

                    e_path = self.current_path + '/{}.xlsx'.format(new_file_name)
                    workbook.save(e_path)
                    workbook.close()
                    arg_parent.switch_frame(Menu)

        ttk.Button(self, text="Sélectionnez le répertoire d'export", command=lambda: on_selection()).grid(row=1,
                                                                                                          column=0)
        ttk.Button(self, text="Valider", command=lambda: confirmation()).grid(row=2, column=0)
        ttk.Button(self, text="Retour", command=lambda: arg_parent.switch_frame(Menu)).grid(row=2, column=1)


#########################################################################
# All functions necessary for launch


def initialisation(arg_root_window):
    directory_path = check_directory()
    check_parameters_json_file(directory_path)
    check_students_csv_file(directory_path)
    check_icon(arg_root_window)


def check_directory(): # Every generated files will be put in a single directory.
    target_path = os.getcwd() + "\\dossier de l'application"
    if not os.path.exists(target_path):
        os.makedirs(target_path)
    return target_path


def check_parameters_json_file(path):
    file_path = path + "\\paramètres validation.json"
    if not os.path.exists(file_path):  # create a json file if not exists.
        with open(file_path, 'w') as parameters_file:
            data = {"actions": [], "departments": []}
            json.dump(data, parameters_file, indent=2)


def check_students_csv_file(path):  # Will initiate basic attributes of students.
    file_path = path + "\\base de données élèves.csv"
    if not os.path.exists(file_path):
        with open(file_path, 'w', newline='') as students_file:
            data = [['First Name', 'last Name', "Promo", 'Student Number']]
            csv_writer = csv.writer(students_file)
            csv_writer.writerows(data)


def check_icon(arg_root_window):
    icon_path = os.getcwd() + "\\icon.png"
    if os.path.exists(icon_path):
        img = Image.open(icon_path)
        img = img.resize((16, 16))
        tk_img = ImageTk.PhotoImage(img)
        arg_root_window.iconphoto(True, tk_img)

#########################################################################


def read_validation_parameters():
    path = os.getcwd() + "\\dossier de l'application\\paramètres validation.json"
    list_actions = []  # List of possible actions with the corresponding department
    list_department = []  # List of department to validate and their threshold
    with open(path, 'r') as json_file:
        data_read = json.load(json_file)

        actions = data_read.get("actions")
        departments = data_read.get("departments")

        for action in actions:
            list_actions.append((action.get("nom"), action.get("pôle")))

        for department in departments:
            list_department.append((department.get("nom"), department.get("seuil")))

    return list_actions, list_department


def generate_text(list_actions, list_departments):
    text_action = "Actions possibles : \n\n"
    for i in range(0, len(list_actions)):
        text_action += "{} dans le pôle {}\n".format(list_actions[i][0], list_actions[i][1])

    text_department = "Pôles à valider : \n\n"
    for j in range(0, len(list_departments)):
        text_department += ("pôle {} avec un nombre de point nécessaire de {}\n"
                        .format(list_departments[j][0], list_departments[j][1]))

    return text_action, text_department


root_window = Tk()
root_window.geometry("400x300")

initialisation(root_window)

App()
root_window.mainloop()

